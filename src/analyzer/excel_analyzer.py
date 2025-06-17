import pandas as pd
import numpy as np
import re
from src.utils.logging_config import get_logger
from collections import defaultdict

class ExcelAnalyzer:
    def __init__(self, file_path):
        """Initialize the Excel analyzer with the file path"""
        self.file_path = file_path
        self.excel_file = None
        self.sheet_names = []
        self.sheet_data = {}
        self.numerical_columns = {}
        self.potential_queries = {}
        self.logger = get_logger(__name__)
        # Define columns that should not be rounded (full account names in CAReportName)
        self.exception_columns = [
            "Margin %",
            "Bad debt % of net revenue",
            "Clinical OT wages % of clinical wages",
            "Contractual adjustment percentage",
            "Payroll tax % of wages"
        ]
        
    
    def load_excel(self):
        """Load the Excel file and extract sheet names.

        This now unmerges all cells before handing the workbook off to
        ``pandas``.  When merged cells are encountered only the top-left
        cell retains its value while the remaining cells are cleared.  This
        prevents accidental duplicate columns when merged headers span
        multiple columns.
        """
        try:
            from openpyxl import load_workbook
            import io

            wb = load_workbook(self.file_path)
            for ws in wb.worksheets:
                for merged in list(ws.merged_cells.ranges):
                    min_row, min_col, max_row, max_col = (
                        merged.min_row,
                        merged.min_col,
                        merged.max_row,
                        merged.max_col,
                    )
                    value = ws.cell(row=min_row, column=min_col).value
                    ws.unmerge_cells(str(merged))
                    # Preserve the original value only in the first cell;
                    # clear the others so they can be pruned later if empty.
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            if r == min_row and c == min_col:
                                ws.cell(row=r, column=c).value = value
                            else:
                                ws.cell(row=r, column=c).value = None

            temp_buffer = io.BytesIO()
            wb.save(temp_buffer)
            temp_buffer.seek(0)

            self.excel_file = pd.ExcelFile(temp_buffer)
            self.sheet_names = self.excel_file.sheet_names
            self.logger.info(
                f"Successfully loaded Excel file with {len(self.sheet_names)} sheets"
            )
            return True
        except Exception as e:
            self.logger.error(f"Failed to load Excel file: {str(e)}")
            return False
    
    def analyze_sheet(self, sheet_name, header_rows=5):
        """Analyze a specific sheet to identify structure and data columns"""
        if not self.excel_file:
            self.logger.error("Excel file not loaded")
            return None
        
        try:
            # Read the sheet with no header initially to analyze structure
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=None)
            
            # Set the sheet name on the DataFrame so methods can access it
            df.name = sheet_name
            
            # First analyze sheet structure
            structure = self._analyze_sheet_structure(df)
            header_indexes = self._detect_headers(df, header_rows)
            
            # Store sheet data
            self.sheet_data[sheet_name] = {
                "dataframe": df,
                "shape": df.shape,
                "structure": structure,
                "header_indexes": header_indexes
            }
            
            # Now detect data area
            data_area = self._detect_data_area(df)
            self.sheet_data[sheet_name]["data_area"] = data_area
            
            # Try to detect columns that are numerical and might contain financial data
            self.numerical_columns[sheet_name] = self._detect_numerical_columns(df)
            
            # Generate potential SQL queries based on sheet analysis
            self.potential_queries[sheet_name] = self._generate_potential_queries(sheet_name)
            
            self.logger.info(f"Analyzed sheet: {sheet_name}")
            return self.sheet_data[sheet_name]
            
        except Exception as e:
            self.logger.error(f"Failed to analyze sheet {sheet_name}: {str(e)}")
            return None
    
    def _analyze_sheet_structure(self, df):
        """Analyze the structure of the sheet to identify headers, data areas, etc."""
        structure = {
            "empty_rows": [],
            "potential_headers": [],
            "data_regions": [],
            "footer_rows": []
        }
        
        # Identify empty rows
        for i, row in df.iterrows():
            if row.isna().all() or (row.astype(str).str.strip().str.len() == 0).all():
                structure["empty_rows"].append(i)
        
        # Identify potential headers (rows with string values and few numbers)
        for i, row in df.iterrows():
            if i in structure["empty_rows"]:
                continue
            
            numeric_count = row.apply(lambda x: isinstance(x, (int, float)) and not np.isnan(x)).sum()
            string_count = row.apply(lambda x: isinstance(x, str)).sum()
            
            if string_count > numeric_count and string_count > 0:
                structure["potential_headers"].append(i)
        
        # Identify data regions (consecutive rows with similar structure)
        current_region = []
        for i in range(df.shape[0]):
            if i in structure["empty_rows"]:
                if current_region:
                    structure["data_regions"].append(current_region)
                    current_region = []
            else:
                current_region.append(i)
        
        if current_region:
            structure["data_regions"].append(current_region)
        
        return structure
    
    def _detect_headers(self, df, max_header_rows=5):
        """Detect likely header rows in the sheet"""
        headers = []
        header_patterns = [
            r"(?i)(total|sum|subtotal|balance|revenue|expense|income|cost|profit|loss)",
            r"(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)",
            r"(?i)(quarter|q1|q2|q3|q4|ytd|mtd)",
            r"(?i)(actual|budget|variance|forecast|plan|target)"
        ]
        
        # Check first few rows for header-like content
        for i in range(min(max_header_rows, df.shape[0])):
            row = df.iloc[i]
            header_matches = 0
            
            for cell in row:
                if not isinstance(cell, str):
                    continue
                    
                for pattern in header_patterns:
                    if re.search(pattern, cell):
                        header_matches += 1
                        break
            
            # If multiple header-like terms found, consider it a header
            if header_matches >= 2:
                headers.append(i)
        
        return headers
    
    def _detect_data_area(self, df):
        """Identify the main data area in the sheet"""
        # Get sheet name from the DataFrame if it exists
        sheet_name = getattr(df, 'name', None)
        
        # If sheet_name is not available, we can't access the structure info properly
        if not sheet_name or sheet_name not in self.sheet_data:
            # Default approach: find first and last non-empty rows
            empty_rows = []
            for i, row in df.iterrows():
                if row.isna().all() or (row.astype(str).str.strip().str.len() == 0).all():
                    empty_rows.append(i)
                    
            start_row = 0
            end_row = df.shape[0] - 1
            
            # Adjust the start row (skip empty rows at the beginning)
            while start_row < df.shape[0] and start_row in empty_rows:
                start_row += 1
                
            # Adjust the end row (skip empty rows at the end)
            while end_row >= 0 and end_row in empty_rows:
                end_row -= 1
                
            return (start_row, end_row)
        
        # If we have structure info, use it
        empty_rows = self.sheet_data[sheet_name].get("structure", {}).get("empty_rows", [])
        potential_headers = self.sheet_data[sheet_name].get("structure", {}).get("potential_headers", [])
        
        # Find first non-empty row after headers
        start_row = 0
        if potential_headers:
            start_row = max(potential_headers) + 1
        
        # Find last non-empty row
        end_row = df.shape[0] - 1
        for i in range(df.shape[0] - 1, -1, -1):
            if i not in empty_rows:
                end_row = i
                break
        
        return (start_row, end_row)
    
    def _detect_numerical_columns(self, df):
        """Detect columns that contain numerical financial data"""
        numerical_cols = {}
        
        # Get sheet name from the DataFrame
        sheet_name = getattr(df, 'name', None)
        if not sheet_name or sheet_name not in self.sheet_data:
            return numerical_cols
        
        # First, identify header row
        header_row = 0
        if self.sheet_data[sheet_name].get("header_indexes"):
            header_row = max(self.sheet_data[sheet_name]["header_indexes"])
        
        # For each column, check if it's mostly numerical
        for col in range(df.shape[1]):
            col_values = df.iloc[header_row+1:, col].dropna()
            if len(col_values) == 0:
                continue
                
            numeric_count = col_values.apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x)).sum()
            numeric_ratio = numeric_count / len(col_values) if len(col_values) > 0 else 0
            
            if numeric_ratio > 0.7:  # If over 70% of values are numeric
                col_header = df.iloc[header_row, col]
                numerical_cols[col] = {
                    "header": col_header,
                    "numeric_ratio": numeric_ratio,
                    "sample_values": col_values.head(3).tolist()
                }
        
        return numerical_cols
    
    def _generate_potential_queries(self, sheet_name):
        """Generate potential SQL queries based on sheet analysis"""
        # This is a simplified version - in reality, this would be much more sophisticated
        potential_queries = []
        
        # Check if sheet exists in our data
        if sheet_name not in self.sheet_data:
            self.logger.warning(f"Sheet {sheet_name} has not been analyzed")
            return potential_queries
        
        # Get column headers if available
        header_indexes = self.sheet_data[sheet_name]["header_indexes"]
        if not header_indexes:
            return potential_queries
            
        header_row = max(header_indexes)
        df = self.sheet_data[sheet_name]["dataframe"]
        
        # Extract headers
        headers = [str(h).strip() for h in df.iloc[header_row] if not pd.isna(h)]
        
        # Look for dimensions (non-numeric columns) that might be in a SQL table
        dimensions = []
        if sheet_name in self.numerical_columns:
            for col, col_info in self.numerical_columns[sheet_name].items():
                header = col_info.get("header")
                numeric_ratio = col_info.get("numeric_ratio", 0)
                if header and numeric_ratio < 0.5:  # Less than 50% numeric = dimension
                    dimensions.append(header)
        
        # Look for metrics (numeric columns)
        metrics = []
        if sheet_name in self.numerical_columns:
            for col, col_info in self.numerical_columns[sheet_name].items():
                header = col_info.get("header")
                numeric_ratio = col_info.get("numeric_ratio", 0)
                if header and numeric_ratio > 0.7:  # More than 70% numeric = metric
                    metrics.append(header)
        
        # Generate a basic query based on sheet name and columns
        if dimensions and metrics:
            # Clean sheet name to make it a potential table name
            table_name = re.sub(r'[^a-zA-Z0-9]', '_', sheet_name)
            
            # Build query
            query = f"SELECT {', '.join(dimensions)},\n"
            query += f"       {', '.join(metrics)}\n"
            query += f"FROM {table_name}"
            
            potential_queries.append({
                "name": f"Basic query for {sheet_name}",
                "sql": query,
                "dimensions": dimensions,
                "metrics": metrics
            })
        
        return potential_queries
    
    def analyze_all_sheets(self):
        """Analyze all sheets in the Excel file"""
        if not self.excel_file:
            self.logger.error("Excel file not loaded")
            return False
            
        for sheet_name in self.sheet_names:
            self.analyze_sheet(sheet_name)
            
        self.logger.info(f"Analyzed all {len(self.sheet_names)} sheets")
        return True
    
    def get_sheet_summary(self, sheet_name=None):
        """Get a summary of the analyzed sheet(s)"""
        if sheet_name:
            if sheet_name not in self.sheet_data:
                self.logger.warning(f"Sheet {sheet_name} has not been analyzed")
                return None
            return {
                "name": sheet_name,
                "shape": self.sheet_data[sheet_name]["shape"],
                "numerical_columns": len(self.numerical_columns.get(sheet_name, {})),
                "potential_queries": len(self.potential_queries.get(sheet_name, [])),
            }
        else:
            # Return summary for all analyzed sheets
            return {
                sheet: {
                    "shape": data["shape"],
                    "numerical_columns": len(self.numerical_columns.get(sheet, {})),
                    "potential_queries": len(self.potential_queries.get(sheet, [])),
                }
                for sheet, data in self.sheet_data.items()
            }
    
    def get_smart_query_suggestions(self, sheet_name):
        """Generate intelligent SQL query suggestions based on sheet content"""
        if sheet_name not in self.sheet_data:
            self.logger.warning(f"Sheet {sheet_name} has not been analyzed")
            return []
            
        # In a real implementation, this would be much more sophisticated
        # using machine learning models to understand the sheet's purpose
        return self.potential_queries.get(sheet_name, [])
    
    def clean_excel_data(self, df, sheet_name):
        """Clean Excel data by rounding all numeric values (except CAReportName) to 2 decimal places for every row."""
        try:
            cleaned_df = df.copy()

            # Try to find the CAReportName column (case-insensitive)
            ca_col = None
            for col in cleaned_df.columns:
                if str(col).strip().lower() == 'careportname':
                    ca_col = col
                    break
            if ca_col is None:
                # Fallback: assume first column is the account name
                ca_col = cleaned_df.columns[0]
                self.logger.warning(f"CAReportName column not found. Defaulting to first column: {ca_col}")

            # Round all numeric columns (except CAReportName) to 2 decimal places for every row
            for idx, row in cleaned_df.iterrows():
                for col in cleaned_df.columns:
                    if col == ca_col:
                        continue
                    val = row[col]
                    if pd.notnull(val) and isinstance(val, (int, float)):
                        cleaned_df.at[idx, col] = round(float(val), 2)

            return cleaned_df
        except Exception as e:
            self.logger.error(f"Error cleaning Excel data: {str(e)}")
            return df  # Return original DataFrame if cleaning fails 
