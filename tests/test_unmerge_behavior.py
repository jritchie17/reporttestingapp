import unittest
from tempfile import NamedTemporaryFile

import pandas as pd

class TestUnmergeBehavior(unittest.TestCase):
    def test_unmerge_clears_other_cells(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")

        from src.analyzer.excel_analyzer import ExcelAnalyzer

        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws["A1"] = "Header"

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            analyzer = ExcelAnalyzer(tmp.name)
            self.assertTrue(analyzer.load_excel())
            df = pd.read_excel(analyzer.excel_file, sheet_name=0, header=None)
            self.assertEqual(df.iloc[0, 0], "Header")
            self.assertTrue(pd.isna(df.iloc[0, 1]))
            self.assertTrue(pd.isna(df.iloc[0, 2]))


if __name__ == "__main__":
    unittest.main()
